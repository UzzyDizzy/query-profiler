# ==========================================================
# scripts/per_method_processors/instance_grid_maps.py
# ==========================================================

from pathlib import Path
import pandas as pd
import numpy as np

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill,
    Border,
    Side
)

from openpyxl.formatting.rule import (
    ColorScaleRule
)


# ==========================================================
# colors
# ==========================================================

PINK=PatternFill(
    fill_type="solid",
    fgColor="FFC0CB"
)

YELLOW=PatternFill(
    fill_type="solid",
    fgColor="FFFF00"
)

BLUE=PatternFill(
    fill_type="solid",
    fgColor="87CEFA"
)

RED=Side(
    style="thick",
    color="FF0000"
)

RED_BORDER=Border(
    left=RED,
    right=RED,
    top=RED,
    bottom=RED
)


# ==========================================================

def run(results_dir):

    results_dir=Path(results_dir)

    csv=results_dir/"ground_truth.csv"

    if not csv.exists():
        return

    df=pd.read_csv(csv)

    outdir=results_dir/"grid_maps"

    outdir.mkdir(
        parents=True,
        exist_ok=True
    )


    # ======================================================
    # representative coordinates
    # ======================================================

    xvals=np.sort(
        df["x1"].unique()
    )

    yvals=np.sort(
        df["x2"].unique()
    )


    lookup={

        (r["x1"],r["x2"]):r

        for _,r in df.iterrows()

    }


    # ======================================================
    # display boundary regions
    # ======================================================

    x_boundary_idx={

        0,
        1,
        len(xvals)-1
    }

    y_boundary_idx={

        0,
        1,
        len(yvals)-1
    }


    def is_boundary(i,j):

        return (

            i in y_boundary_idx
            or
            j in x_boundary_idx
        )


    # ======================================================

    def create_sheet(

        outfile,
        value_col,
        axis,
        fill_color=None,
        use_qerr=False,
        joint=False

    ):

        wb=Workbook()

        ws=wb.active


        # ----------------------------
        # headers
        # ----------------------------

        ws.cell(
            row=1,
            column=1,
            value=""
        )

        for j,x in enumerate(
                xvals,
                start=2
        ):

            ws.cell(
                row=1,
                column=j,
                value=round(x,6)
            )

        for i,y in enumerate(
                yvals,
                start=2
        ):

            ws.cell(
                row=i,
                column=1,
                value=round(y,6)
            )


        # ----------------------------
        # cells
        # ----------------------------

        qerr_cells=[]

        for i,y in enumerate(
                yvals
        ):

            for j,x in enumerate(
                    xvals
            ):

                r=lookup[
                    (x,y)
                ]

                val=r.get(
                    value_col,
                    np.nan
                )

                cell=ws.cell(
                    row=i+2,
                    column=j+2,
                    value=val
                )


                boundary=is_boundary(
                    i,
                    j
                )


                # ====================
                # boundary pink
                # ====================

                if boundary:

                    cell.fill=PINK


                # ====================
                # yellow / blue
                # ====================

                elif (

                    fill_color is not None
                    and
                    pd.notna(val)
                    and
                    val!=0

                ):

                    cell.fill=fill_color


                # ====================
                # red border
                # ====================

                if (

                    joint
                    and
                    not boundary

                ):

                    v1=r.get(
                        f"dS_x1_axis{axis}",
                        0
                    )

                    v2=r.get(
                        f"dS_x2_axis{axis}",
                        0
                    )

                    both=(

                        pd.notna(v1)
                        and
                        pd.notna(v2)
                        and
                        v1!=0
                        and
                        v2!=0
                    )

                    if both:

                        cell.border=RED_BORDER


                # ====================
                # collect qerr cells
                # ====================

                if (

                    use_qerr
                    and
                    not boundary

                ):

                    q=(
                        r.get(
                            f"adjacent_qerr_x{axis}",
                            np.nan
                        )
                    )

                    if pd.notna(q):

                        cell.value=q

                        qerr_cells.append(
                            cell.coordinate
                        )


        # ==========================
        # qerr color scale
        # ==========================

        if len(qerr_cells)>0:

            rng=" ".join(
                qerr_cells
            )

            ws.conditional_formatting.add(

                rng,

                ColorScaleRule(

                    start_type="min",
                    start_color="FFFFFF",

                    end_type="max",
                    end_color="006400"

                )
            )


        wb.save(
            outfile
        )


    # ======================================================
    # axis1
    # ======================================================

    create_sheet(
        outdir/"axis1_dS_x1.xlsx",
        "dS_x1_axis1",
        axis=1,
        fill_color=YELLOW
    )

    create_sheet(
        outdir/"axis1_dS_x2.xlsx",
        "dS_x2_axis1",
        axis=1,
        fill_color=BLUE
    )

    create_sheet(
        outdir/"axis1_joint_dS.xlsx",
        "joint_dS_axis1",
        axis=1,
        use_qerr=True,
        joint=True
    )


    # ======================================================
    # axis2
    # ======================================================

    create_sheet(
        outdir/"axis2_dS_x1.xlsx",
        "dS_x1_axis2",
        axis=2,
        fill_color=YELLOW
    )

    create_sheet(
        outdir/"axis2_dS_x2.xlsx",
        "dS_x2_axis2",
        axis=2,
        fill_color=BLUE
    )

    create_sheet(
        outdir/"axis2_joint_dS.xlsx",
        "joint_dS_axis2",
        axis=2,
        use_qerr=True,
        joint=True
    )


    print(
        f"saved: {outdir}"
    )